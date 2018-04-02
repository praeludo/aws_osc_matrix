import requests
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

import os
import re
import string
from pathlib import Path
from io import BytesIO
import subprocess

from resources import *

def get_or_dl_pdf(url):
    filename = url.split('/')[-1]
    tmp_path = Path(os.path.join(os.sep, 'tmp', filename))
    if not tmp_path.exists():
        print(f"Downloading documentation from {url}... ", end='', flush=True)
        pdf_file = requests.get(url).content
        print("PDF downloaded... ", end='')
        tmp_path.write_bytes(pdf_file)
    else:
        print(f"Opening {tmp_path}... ", end='', flush=True)
        pdf_file = tmp_path.read_bytes()
    return(BytesIO(pdf_file))

def extract_from_awspdf(url):
    pdf_file = get_or_dl_pdf(url)
    print("Analyzing pdf... ", end='', flush=True),
    document = PDFDocument(PDFParser(pdf_file))
    outlines = document.get_outlines()
    current_section = ""
    actions = []
    for level, title, *_ in outlines:
        if level == 1:
            current_section = title
        if current_section == "Actions" and level == 2:
            actions.append(title)
    return actions

def extract_from_oscdoc(url):
    actions = []
    osc_soup = BeautifulSoup(requests.get(url).content, 'html.parser')
    print(f'Analyzing {url}... ', end='', flush=True)
    for h2 in osc_soup.find_all('h2'):
        if h2.string == 'Resources':
            for command in h2.parent.find_all('h4'):
                actions.append(command.string)
    return actions

def extract_from_oscwiki(url):
    if 'osc_commands' not in globals():
        global osc_commands
        print(f"Scrapping {url}... ")
        aws_osc = {'API EC2 (2016-09-15)': 'ec2',
                   'API IAM (2010-05-08)': 'iam',
                   'API ELB (2012-06-01)': 'elb',
                   'API Direct Connect (2012-10-25)': 'dircon'}
        osc_commands = {}
        osc_soup = BeautifulSoup(requests.get(url).content, 'html.parser')
        osc_tab = osc_soup.find('table')
        for tr in osc_tab.find_all('tr'):
            command_details = [child.string for child in tr.children]
            try:
                if aws_osc[command_details[3]] in osc_commands.keys():
                    osc_commands[aws_osc[command_details[3]]].append(command_details[2])
                else:
                    osc_commands[aws_osc[command_details[3]]] = [command_details[2]]
            except KeyError:
                pass
    return osc_commands[service]

def write_to_excel(commands):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for service in commands:
        ws = wb.create_sheet(service)
        ws.title = service
        ws['A1'] = "Command"
        i = 1
        columns = {}
        for provider in list(commands[service]):
            columns[provider] = string.ascii_letters[i].upper()
            ws[f'{columns[provider]}1'] = provider
            i += 1

        line = 2        
        for command in (sorted({command 
                                for provider in commands[service].keys()
                                for command in commands[service][provider]})):
            ws[f'A{line}'] = command
            for provider in columns:
                if command in commands[service][provider]:
                    ws[f'{columns[provider]}{line}'] = 'X'
            line += 1
        line -= 1
        tab = Table(displayName=f"{service}_table", ref=f'A1:{columns[provider]}{line}')
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        line += 1
        for provider, column in columns.items():
            val = f'=NBVAL({service}_table[[{provider}]])'
            ws[f'{column}{line}'].value = f'=NBVAL({service}_table[[{provider}]])'
        ws.add_table(tab)
    wb.save(SAVE_TO)
    print(f"File saved to {SAVE_TO}")
    subprocess.run(('open', SAVE_TO))

if __name__ == "__main__":
    providers = [key.split('_')[0] for key in locals().keys()
             if re.match('(\w+)_RESOURCES', key)]
    services = set(key for provider in providers 
                    for key in eval('%s_RESOURCES' % provider))
    commands = {service.lower(): {} for service in services}

    for provider in providers:
        print(provider)
        for service, url in eval('%s_RESOURCES' % provider).items():
            actions =  eval('extract_from_%s("%s")' % (provider.lower(), url))
            print(f"{service.upper()}: ", end=''),
            print(f"{len(actions)} actions extracted")
            commands[service][provider.lower()] = actions
        print(' ')

    write_to_excel(commands)
